from fastapi import FastAPI, Body, Response, Cookie
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import RedirectResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles

from openpyxl import load_workbook
from openpyxl import Workbook

app = FastAPI()

#app.mount("/static", StaticFiles(directory="public", html=True))
#app.mount("/static", StaticFiles(directory="/", html=True))



link = "https://6438-176-15-250-9.ngrok-free.app/"

origins = [
    link#,
    #"https://127.0.0.1:8000",
    #"https://localhost:8000"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST"],#, "OPTIONS", "DELETE", "PATH", "PUT"],
    allow_headers=["Content-Type", "Accept", "Location", "Allow", "Content-Disposition", "Sec-Fetch-Dest"],
)



app.mount("/static", StaticFiles(directory="public", html=True))
app.mount("/static", StaticFiles(directory="/", html=True))



@app.get("/", response_class=HTMLResponse)
def root():
    return RedirectResponse(f"{link}static/menu.html")



@app.get("/menu")
def menu():
    wb = load_workbook('list.xlsx')
    sheet = wb['all']

    j=0
    HEADERS = {j : "ОГЛАВЛЕНИЕ"}
    for i in range(2, sheet.max_row+1):
        header = sheet[f"A{i}"].value
        if header not in HEADERS.values():
            j+=1
            HEADERS[j] = header
    
    return HEADERS



@app.get("/menu_plus")
def menu_plus():
    wb = load_workbook('list.xlsx')
    sheet = wb['all']

    j=0
    HEADERS = []
    for i in range(2, 98):
        header = sheet[f"A{i}"].value
        if header not in HEADERS:
            j+=1
            HEADERS.append(header)

    MENU = dict()
    for header in HEADERS:
        MENU[header] = []
        for i in range(2, 98):
            hdr = sheet[f"A{i}"].value
            if (hdr != None) and (hdr == header):# or (hdr.lower() in header.lower()):
                card = {
                    "id" : i,
                    "name" : sheet[f"A{i}"].value,
                    "text" : sheet[f"B{i}"].value,
                    "img" : sheet[f"C{i}"].value
                }
                MENU[header].append(card)
    
    return MENU

    

@app.get("/search_by_name/{name}")
def byname(name):
    name = str(name)
    wb = load_workbook('list.xlsx')
    sheet = wb['all']

    CARDS = []
    for i in range(2, sheet.max_row+1):
        nm = sheet[f"A{i}"].value
        if (nm != None) and ((nm == name) or (name.lower() in nm.lower())):
            card = {
                "id" : i,
                "name" : sheet[f"A{i}"].value,
                "text" : sheet[f"B{i}"].value,
                "img" : sheet[f"C{i}"].value
            }
            CARDS.append(card)
    
    return CARDS


def in_tags(s):
    open_tags = []
    frm = 0
    while s.find("<", frm) != -1:
        inx = s.find("<", frm)
        open_tags.append(inx)
        frm = inx + 1

    close_tags = []
    frm=0
    while s.find(">", frm) != -1:
        inx = s.find(">", frm)
        close_tags.append(inx)
        frm = inx + 1
    
    list_tag_index = []
    for i in range(len(open_tags)):
	    for j in range(open_tags[i], close_tags[i]+1):
		    list_tag_index.append(j)
    
    if s.find("<style>") != -1:
        for j in range(s.find("<style>"), s.find("</style>")+1):
            if j not in list_tag_index:
                list_tag_index.append(j)

    return list_tag_index

def add_tag(strg, adound):
    s = strg.lower()
    a = adound.lower()

    index = []
    frm = 0
    while s.find(a, frm) != -1:
        inx = s.find(a, frm)
        index.append(inx)
        frm = inx + len(a)

    ban = in_tags(strg)
    k=0
    for inx in index:
        if inx not in ban:
            i = inx + k
            strg = strg[: i] + "<span class =\'target\'>" + strg[i : i+len(a)] + "</span>" + strg[i+len(a) :]
            k+=len("<span class =\'target\'>" + "</span>")

    return strg

def ban(s):
    banned = '''
    <style>table.iksweb{
	width: 100%;
	border-collapse:collapse;
	border-spacing:0;
	height: auto;
}
table.iksweb,table.iksweb td, table.iksweb th {
	border: 1px solid #595959;
}
table.iksweb td,table.iksweb th {
	padding: 3px;
	width: 30px;
	height: 35px;
}
table.iksweb th {
	background: #347c99; 
	color: #fff; 
	font-weight: normal;
}</style><table class="iksweb">
	<tbody><tr><th>

    </th><th>
    </th></tr>

    </td>
		</tr>
	</tbody>
</table>
    '''
    if s in banned:
        return True
    else:
        return False


@app.get("/search_by_text/{text}")
def bytext(text):
    text = str(text)
    wb = load_workbook('list.xlsx')
    sheet = wb['all']

    CARDS = []
    start_i = []
    for i in range(2, sheet.max_row+1):
        nm = sheet[f"A{i}"].value
        txt = sheet[f"B{i}"].value
        if (txt != None) and ((sheet[f"A{i}"].value == text) or (text.lower() in nm.lower()) or (text.lower() in txt.lower())) and (not ban(text)):
            Mesg = sheet[f"B{i}"].value
            
            Mesg = add_tag(Mesg, text)

            card = {
                "id" : i,
                "name" : sheet[f"A{i}"].value,
                "text" : Mesg,
                "img" : sheet[f"C{i}"].value
            }
            CARDS.append(card)
    
    return CARDS

@app.get("/application")
def application():
    text = "ПРИЛОЖЕНИЕ"
    wb = load_workbook('list.xlsx')
    sheet = wb['all']

    CARDS = []
    for i in range(2, sheet.max_row+1):
        nm = sheet[f"A{i}"].value
        if (nm != None) and ((sheet[f"A{i}"].value == "СОКРАЩЕНИЯ") or (sheet[f"A{i}"].value == text) or (text.lower() in nm.lower())):
            card = {
                "id" : i,
                "name" : sheet[f"A{i}"].value,
                "text" : sheet[f"B{i}"].value,
            }
            CARDS.append(card)
    
    return CARDS

#в конце
@app.get("img/{file}")
def files(file: str):
    #/imgs/1_1.jpg
    file = "/" + file
    return FileResponse(file) 